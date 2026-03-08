"""
Document ingestion pipeline for agente-enam-const.

Collections (in RAG priority order):
  1. constituicao        - Constituição Federal (downloaded from planalto.gov.br)
  2. repercussao_geral   - Repercussão Geral mérito julgado (XLSX)
  3. sumulas_vinculantes - Súmulas vinculantes do STF (JSON)
  4. doutrina            - DIREITO CONSTITUCIONAL.pdf
  5. informativos        - Informativos temáticos 2024/2025/2026 (DOCX)
"""

import os
import re
import json
import hashlib
import requests
import chromadb

from bs4 import BeautifulSoup
from pypdf import PdfReader
from docx import Document
import openpyxl
from chromadb.utils.embedding_functions import SentenceTransformerEmbeddingFunction

BASE_DIR = os.path.dirname(__file__)
DATA_DIR = os.path.join(BASE_DIR, "data")
CHROMA_DIR = os.path.join(DATA_DIR, "chroma_db")

EMBEDDING_MODEL = "paraphrase-multilingual-mpnet-base-v2"

CONSTITUICAO_URL = "https://www.planalto.gov.br/ccivil_03/constituicao/constituicao.htm"
CONSTITUICAO_CACHE = os.path.join(DATA_DIR, "constituicao.html")


def get_chroma_client():
    os.makedirs(CHROMA_DIR, exist_ok=True)
    return chromadb.PersistentClient(path=CHROMA_DIR)


def get_embedding_fn():
    return SentenceTransformerEmbeddingFunction(model_name=EMBEDDING_MODEL)


def _make_id(text: str, prefix: str = "") -> str:
    digest = hashlib.md5(text.encode("utf-8")).hexdigest()
    return f"{prefix}{digest}"


def _chunk_text(text: str, chunk_size: int = 800, overlap: int = 100) -> list[str]:
    text = text.strip()
    if not text:
        return []
    chunks = []
    start = 0
    while start < len(text):
        end = start + chunk_size
        chunks.append(text[start:end])
        start += chunk_size - overlap
    return chunks


# ─────────────────────────────────────────────
# 1. Constituição Federal
# ─────────────────────────────────────────────

def _download_constituicao() -> str:
    os.makedirs(DATA_DIR, exist_ok=True)
    if os.path.exists(CONSTITUICAO_CACHE):
        print("  [constituição] Usando cache local.")
        with open(CONSTITUICAO_CACHE, "r", encoding="latin-1") as f:
            return f.read()

    print("  [constituição] Baixando de planalto.gov.br...")
    resp = requests.get(CONSTITUICAO_URL, timeout=60, headers={"User-Agent": "Mozilla/5.0"})
    resp.raise_for_status()
    with open(CONSTITUICAO_CACHE, "w", encoding="latin-1") as f:
        f.write(resp.text)
    return resp.text


def _parse_constituicao(html: str) -> list[dict]:
    """Parse the HTML and return list of {artigo, texto} dicts."""
    soup = BeautifulSoup(html, "lxml")
    text = soup.get_text(separator="\n")

    # Split by Artigo pattern
    artigo_pattern = re.compile(r"(Art\.\s*\d+[º°]?\s*[-–]?\s*.+?)(?=Art\.\s*\d+[º°]?\s*[-–]?|\Z)", re.DOTALL)
    matches = artigo_pattern.findall(text)

    chunks = []
    for match in matches:
        cleaned = re.sub(r"\n{3,}", "\n\n", match.strip())
        if len(cleaned) < 30:
            continue
        chunks.append(cleaned)

    # Also add preamble / ADCT sections that don't start with Art.
    if not chunks:
        # Fallback: generic chunking
        chunks = _chunk_text(text, chunk_size=1000, overlap=150)

    return chunks


def ingest_constituicao(client: chromadb.ClientAPI, ef):
    collection = client.get_or_create_collection("constituicao", embedding_function=ef)
    if collection.count() > 0:
        print(f"  [constituição] Já indexada ({collection.count()} chunks). Pulando.")
        return

    html = _download_constituicao()
    chunks = _parse_constituicao(html)
    print(f"  [constituição] {len(chunks)} chunks extraídos.")

    batch_size = 50
    for i in range(0, len(chunks), batch_size):
        batch = chunks[i:i + batch_size]
        collection.add(
            documents=batch,
            ids=[_make_id(c, "cf_") for c in batch],
            metadatas=[{"source": "Constituição Federal", "chunk": i + j} for j, c in enumerate(batch)],
        )
    print(f"  [constituição] Indexação concluída: {collection.count()} chunks.")


# ─────────────────────────────────────────────
# 2. Repercussão Geral (XLSX)
# ─────────────────────────────────────────────

def ingest_repercussao_geral(client: chromadb.ClientAPI, ef):
    collection = client.get_or_create_collection("repercussao_geral", embedding_function=ef)
    if collection.count() > 0:
        print(f"  [repercussão geral] Já indexada ({collection.count()} chunks). Pulando.")
        return

    xlsx_path = os.path.join(BASE_DIR, "Repercussão Geral - Mérito Julgado  (1).xlsx")
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    # Expected: Número tema, Título tema, Descrição tema, Tese, Processo paradigma,
    #           Ramos do Direito, Assunto, Histórico Processo Paradigma

    docs, ids, metas = [], [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        numero = row_dict.get("Número tema") or ""
        titulo = row_dict.get("Título tema") or ""
        descricao = row_dict.get("Descrição tema") or ""
        tese = row_dict.get("Tese") or ""
        ramo = row_dict.get("Ramos do Direito") or ""

        text = (
            f"Tema {numero}: {titulo}\n"
            f"Descrição: {descricao}\n"
            f"Tese: {tese}\n"
            f"Ramo: {ramo}"
        ).strip()

        if not text or len(text) < 20:
            continue

        docs.append(text)
        ids.append(_make_id(text, "rg_"))
        metas.append({"source": "Repercussão Geral", "tema": str(numero)})

    print(f"  [repercussão geral] {len(docs)} temas extraídos.")

    batch_size = 50
    for i in range(0, len(docs), batch_size):
        collection.add(
            documents=docs[i:i + batch_size],
            ids=ids[i:i + batch_size],
            metadatas=metas[i:i + batch_size],
        )
    print(f"  [repercussão geral] Indexação concluída: {collection.count()} chunks.")


# ─────────────────────────────────────────────
# 3. Súmulas Vinculantes (JSON)
# ─────────────────────────────────────────────

def ingest_sumulas_vinculantes(client: chromadb.ClientAPI, ef):
    collection = client.get_or_create_collection("sumulas_vinculantes", embedding_function=ef)
    if collection.count() > 0:
        print(f"  [súmulas vinculantes] Já indexada ({collection.count()} chunks). Pulando.")
        return

    json_path = os.path.join(BASE_DIR, "Súmulas vinculantes.json")
    with open(json_path, "r", encoding="utf-8") as f:
        sumulas = json.load(f)

    docs, ids, metas = [], [], []
    for s in sumulas:
        numero = s.get("numero", "")
        nome = s.get("nome", "")
        enunciado = s.get("enunciado", "")
        ramo = s.get("ramo_direito", "")

        text = f"{nome}\nEnunciado: {enunciado}\nRamo: {ramo}".strip()
        docs.append(text)
        ids.append(_make_id(text, "sv_"))
        metas.append({"source": "Súmulas Vinculantes", "numero": str(numero)})

    collection.add(documents=docs, ids=ids, metadatas=metas)
    print(f"  [súmulas vinculantes] Indexação concluída: {collection.count()} chunks.")


# ─────────────────────────────────────────────
# 4. Doutrina – DIREITO CONSTITUCIONAL.pdf
# ─────────────────────────────────────────────

def ingest_doutrina(client: chromadb.ClientAPI, ef):
    collection = client.get_or_create_collection("doutrina", embedding_function=ef)
    if collection.count() > 0:
        print(f"  [doutrina] Já indexada ({collection.count()} chunks). Pulando.")
        return

    pdf_path = os.path.join(BASE_DIR, "DIREITO CONSTITUCIONAL.pdf")
    reader = PdfReader(pdf_path)
    print(f"  [doutrina] {len(reader.pages)} páginas encontradas.")

    full_text = "\n".join(
        page.extract_text() or "" for page in reader.pages
    )
    chunks = _chunk_text(full_text, chunk_size=900, overlap=120)
    print(f"  [doutrina] {len(chunks)} chunks gerados.")

    docs, ids, metas = [], [], []
    for i, chunk in enumerate(chunks):
        if len(chunk.strip()) < 40:
            continue
        docs.append(chunk)
        ids.append(_make_id(chunk, "doc_"))
        metas.append({"source": "Doutrina - Direito Constitucional", "chunk": i})

    batch_size = 50
    for i in range(0, len(docs), batch_size):
        collection.add(
            documents=docs[i:i + batch_size],
            ids=ids[i:i + batch_size],
            metadatas=metas[i:i + batch_size],
        )
    print(f"  [doutrina] Indexação concluída: {collection.count()} chunks.")


# ─────────────────────────────────────────────
# 5. Informativos Temáticos (DOCX)
# ─────────────────────────────────────────────

def _parse_docx(path: str) -> str:
    doc = Document(path)
    paragraphs = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(paragraphs)


def ingest_informativos(client: chromadb.ClientAPI, ef):
    collection = client.get_or_create_collection("informativos", embedding_function=ef)
    if collection.count() > 0:
        print(f"  [informativos] Já indexada ({collection.count()} chunks). Pulando.")
        return

    docx_files = [
        ("Informativo_tematico_2024_atualizado_1163.docx", "Informativo Temático 2024"),
        ("Informativo_tematico_2025_atualizado_1202.docx", "Informativo Temático 2025"),
        ("Informativo_tematico_2026_atualizado_1205.docx", "Informativo Temático 2026"),
    ]

    all_docs, all_ids, all_metas = [], [], []

    for filename, label in docx_files:
        path = os.path.join(BASE_DIR, filename)
        if not os.path.exists(path):
            print(f"  [informativos] Arquivo não encontrado: {filename}")
            continue

        text = _parse_docx(path)
        chunks = _chunk_text(text, chunk_size=850, overlap=100)
        print(f"  [informativos] {label}: {len(chunks)} chunks.")

        for i, chunk in enumerate(chunks):
            if len(chunk.strip()) < 40:
                continue
            all_docs.append(chunk)
            all_ids.append(_make_id(chunk, "inf_"))
            all_metas.append({"source": label, "chunk": i})

    batch_size = 50
    for i in range(0, len(all_docs), batch_size):
        collection.add(
            documents=all_docs[i:i + batch_size],
            ids=all_ids[i:i + batch_size],
            metadatas=all_metas[i:i + batch_size],
        )
    print(f"  [informativos] Indexação concluída: {collection.count()} chunks.")


# ─────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────

def run_ingestion():
    print("=== Iniciando ingestão de documentos ===\n")
    client = get_chroma_client()
    ef = get_embedding_fn()

    print("[1/5] Constituição Federal")
    ingest_constituicao(client, ef)

    print("\n[2/5] Repercussão Geral")
    ingest_repercussao_geral(client, ef)

    print("\n[3/5] Súmulas Vinculantes")
    ingest_sumulas_vinculantes(client, ef)

    print("\n[4/5] Doutrina (PDF)")
    ingest_doutrina(client, ef)

    print("\n[5/5] Informativos Temáticos")
    ingest_informativos(client, ef)

    print("\n=== Ingestão concluída! ===")


if __name__ == "__main__":
    run_ingestion()
